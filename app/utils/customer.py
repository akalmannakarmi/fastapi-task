from pathlib import Path
from app.db.database import SessionLocal
from app.db.models import Customer
from openpyxl import load_workbook
from typing import Tuple, List, Dict
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session
import csv

REQUIRED_COLUMNS = {"name", "email", "amount"}
BATCH_SIZE = 1000


def validate_file_columns(path: Path) -> None:
    ext = path.suffix.lower()
    if ext == ".csv":
        header = get_csv_header(path)
    elif ext in (".xls", ".xlsx"):
        header = get_excel_header(path)
    else:
        raise ValueError("Unsupported file type")

    validate_header(header)


def get_csv_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames or []


def get_excel_header(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    return list(next(rows, []))


def validate_header(header: list[str] | tuple[str, ...]) -> None:
    if not header:
        raise ValueError("File has no header")

    normalized = {str(col).strip().lower() for col in header if col}
    missing = REQUIRED_COLUMNS - normalized

    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


def batch_and_insert(path: Path, user_id: int) -> Tuple[int, int, int]:
    ext = path.suffix.lower()

    if ext == ".csv":
        return batch_insert_csv(path, user_id)
    elif ext in (".xls", ".xlsx"):
        return batch_insert_excel(path, user_id)
    else:
        raise ValueError("Unsupported file type")


def batch_insert_csv(path: Path, user_id: int) -> Tuple[int, int, int]:
    total = success = failed = 0
    seen_emails = set()
    batch: List[Dict] = []

    with SessionLocal() as db, path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for raw_row in reader:
            total += 1
            row = validate_row(raw_row)

            if not row or row["email"] in seen_emails:
                failed += 1
                continue

            seen_emails.add(row["email"])
            batch.append(row)

            if len(batch) >= BATCH_SIZE:
                ok, bad = batch_insert_customers(db, batch, user_id)
                success += ok
                failed += bad
                batch.clear()

        # flush remaining
        ok, bad = batch_insert_customers(db, batch, user_id)
        success += ok
        failed += bad

    return total, failed, success


def batch_insert_excel(path: Path, user_id: int) -> Tuple[int, int, int]:
    total = success = failed = 0
    seen_emails = set()
    batch: List[Dict] = []

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [h.strip().lower() for h in next(rows)]

    with SessionLocal() as db:
        for values in rows:
            total += 1
            raw_row = dict(zip(headers, values))
            row = validate_row(raw_row)

            if not row or row["email"] in seen_emails:
                failed += 1
                continue

            seen_emails.add(row["email"])
            batch.append(row)

            if len(batch) >= BATCH_SIZE:
                ok, bad = batch_insert_customers(db, batch, user_id)
                success += ok
                failed += bad
                batch.clear()

        ok, bad = batch_insert_customers(db, batch, user_id)
        success += ok
        failed += bad

    return total, failed, success


def validate_row(row: Dict) -> Dict | None:
    try:
        if not REQUIRED_COLUMNS.issubset(row):
            return None

        return {
            "name": str(row["name"]).strip(),
            "email": str(row["email"]).strip().lower(),
            "amount": int(row["amount"]),
        }
    except Exception:
        return None


def batch_insert_customers(
    db: Session,
    rows: List[Dict],
    user_id: int,
) -> Tuple[int, int]:
    if not rows:
        return 0, 0

    stmt = (
        insert(Customer)
        .values([{**row, "user_id": user_id} for row in rows])
        .on_conflict_do_nothing(index_elements=["email", "user_id"])
        .returning(Customer.id)
    )

    result = db.execute(stmt)
    inserted = len(result.fetchall())

    db.commit()

    failed = len(rows) - inserted
    return inserted, failed
